import logging
from time import sleep
from random import choice
from bs4 import BeautifulSoup
from requests import get, exceptions
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

USER_AGENTS_FILE = "resources/user_agents.txt"
PROXIES_FILE = "resources/proxies.txt"
BASE_PAGES_FILE = "resources/base_pages.txt"
COMPARE_PAGES_FILE = "resources/compare_pages.txt"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Scraper:
    def __init__(self, user_agents_file, proxies_file, base_pages_file, compare_pages_file):
        self.user_agents = self.load_data_from_file(user_agents_file)
        self.proxies = self.load_data_from_file(proxies_file)
        self.base_pages = self.load_urls(base_pages_file)
        self.compare_pages = self.load_urls(compare_pages_file)

    def load_data_from_file(self, file_path):
        """
        Load data from a file.

        Args:
            file_path (str): Path to the file.
        Returns:
            list: List of data from the file.
        """

        with open(file_path, 'r') as file:
            return [line.strip() for line in file if line.strip()]

    def load_urls(self, file_path):
        """
        Load URLs, tags, and styles from a file.

        Args:
            file_path (str): Path to the file containing URLs, tags, and styles.
        Returns:
            list: List of tuples containing URL, tag, and style.
        """

        urls_info = []
        with open(file_path, 'r') as file:
            for line in file:
                parts = line.strip().split("|")
                if len(parts) == 3:
                    urls_info.append((parts[0].strip(), parts[1].strip(), parts[2].strip()))
        return urls_info

    def fetch_url(self, url):
        """
        Fetch content from a URL.

        Args:
            url (str): URL to fetch content from.
        Returns:
            str: Content of the URL.
        """

        try:
            headers = {'User-Agent': choice(self.user_agents)}
            if self.proxies:
                proxy = choice(self.proxies)
                proxies = {'http': f'http://{proxy}', 'https': f'http://{proxy}'}
                response = get(url, headers=headers, proxies=proxies)
            else:
                response = get(url, headers=headers)

            response.raise_for_status()
            return response.text
        except exceptions.HTTPError:
            logging.exception(f"Error fetching {url}")
            return f"Error fetching {url}"

    def scrape_content_by_style(self, urls_info):
        """
        Scrape contents based on style from given URLs.

        Args:
            urls_info (list): List of tuples containing URL, tag, and style.
        Returns:
            list: List of contents scraped from the URLs.
        """

        aggregated_contents = []
        for url, tag, style in urls_info:
            logging.info(f"Fetching {url}")
            content = self.fetch_url(url)
            if content:
                soup = BeautifulSoup(content, 'html.parser')
                tags = soup.find_all(tag, style=lambda value: value and style in value)
                aggregated_contents.extend([tag.get_text().strip() for tag in tags])
            sleep(choice(range(1, 5)))
        return aggregated_contents

    def compare_and_export_to_excel(self, list1, list2, excel_file_name):
        """
        Compare two lists and export the result to an Excel file.

        Args:
            list1 (list): First list to compare.
            list2 (list): Second list to compare.
            excel_file_name (str): Name of the Excel file to export to.
        """

        wb = Workbook()
        ws = wb.active

        center_aligned_text = Alignment(horizontal="center", vertical="center")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        ws.append(["Category Items", "Is Exist"])
        for item in list1:
            match = any(item in s for s in list2)
            result = "✓" if match else "✕"
            row = [item, result]
            ws.append(row)

            cell = ws.cell(row=ws.max_row, column=2)
            cell.fill = green_fill if result == "✓" else red_fill
            cell.alignment = center_aligned_text

        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 25

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_aligned_text
        wb.save(excel_file_name)

if __name__ == "__main__":
    scraper = Scraper(USER_AGENTS_FILE, PROXIES_FILE, BASE_PAGES_FILE, COMPARE_PAGES_FILE)

    a_page_content = scraper.scrape_content_by_style(scraper.base_pages)
    b_page_content = scraper.scrape_content_by_style(scraper.compare_pages)
    scraper.compare_and_export_to_excel(a_page_content, b_page_content, "comparison.xlsx")
